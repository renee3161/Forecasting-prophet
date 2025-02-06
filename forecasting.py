# -*- coding: utf-8 -*-
"""
Created on Wed Jan 29 10:29:24 2025

@author: Inventory
"""

import pandas as pd
from prophet import Prophet
import matplotlib.pyplot as plt
from matplotlib import pyplot
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# Read the Excel file containing the sales data
df = pd.read_excel('product_sold_quantities.xlsx')

# Clean the data
df = df.dropna(how='any')  # Drop rows with missing values

# Create a folder to store plots if it doesn't exist
if not os.path.exists('forecast_plots'):
    os.makedirs('forecast_plots')

# Prepare a list to collect forecast results


# Create a new Excel workbook and writer
wb = Workbook()
ws = wb.active
ws.title = "Forecast Results"
ws_plots = wb.create_sheet(title="Forecast Plots")



start_date = pd.to_datetime('2025-01-01')
end_date = pd.to_datetime('2025-12-01')

# Create a list of month-year (e.g. Jan 2025, Feb 2025, ...)
month_year_list = [start_date.strftime('%Y-%m-%d')]  # Start with the first month
while start_date < end_date:
    start_date = start_date + pd.DateOffset(months=1)
    month_year_list.append(start_date.strftime('%Y-%m-%d'))

# Write the headers to the Excel file
# headers = ['Product', 'Date', 'Forecast', 'Lower Bound', 'Upper Bound']
headers = ['Product'] + month_year_list
ws.append(headers)
row_num = 2 
count = 0
# Iterate through each product
for product in df['Product']:
    if count <= 1:
        forecast_results = []
        # Select the data for this product
        product_sales = df[df['Product'] == product].iloc[:, 1:]
    
        # Convert the sales data into long format with columns 'ds' for the date and 'y' for the sales
        sales_data = product_sales.transpose()
        sales_data = sales_data.reset_index()
        sales_data.columns = ['ds', 'y']
        
        sales_data.plot()
        pyplot.show()
        
        # Convert 'ds' to datetime format (month-year format)
        sales_data['ds'] = pd.to_datetime(sales_data['ds'], format='%d/%m/%y')
        
        # Initialize Prophet model
        model = Prophet()
        
        # Optionally add monthly seasonality
        model.add_seasonality(name='monthly', period=30.5, fourier_order=8)
        
        # Fit the model
        model.fit(sales_data)
        
        # Create a dataframe for future dates (next 12 months)
        future = model.make_future_dataframe(periods=12, freq='MS')
        
        # Forecast future sales
        forecast = model.predict(future)
        
        # Add the product name to the forecast
        forecast['Product'] = product
        
        # Append the forecasted results to the list (with product name)
        forecast_results.append(forecast[['Product', 'ds', 'yhat', 'yhat_lower', 'yhat_upper']])
        for forecast in forecast_results:
            for index, row in forecast.iterrows():
              # Write the product name to the first column
                ws.cell(row=row_num, column=1, value=row['Product'])

                # Extract the month-year from the forecasted date
                forecast_date = pd.to_datetime(row['ds']).strftime('%Y-%m-%d')  # Ensure it matches month_year_list format

                # Check if the forecasted date matches any month in the month_year_list
                if forecast_date in month_year_list:
                    col_index = month_year_list.index(forecast_date) + 2
                    ws.cell(row=row_num, column=col_index, value=row['yhat'])

        
        # Plot the forecasted sales (historical and predicted)
        fig = model.plot(forecast)
        plt.title(f"Sales Forecast for {product}")
        
        # Save the plot to the specific folder
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close() 
    
        image_stream.seek(0)  # Rewind the stream to the beginning
        img = Image(image_stream)
    
        # Add the image to the Excel file (create a new sheet for each product's forecast plot)
        row_pos = (row_num - 2) * 20 + 1
        img.width = 600  # Optional: Set image width
        img.height = 400  # Optional: Set image height
        ws_plots.add_image(img, f'G{row_pos}')  # Adjust location where the plot should go
        row_num +=1
# Concatenate all the forecast results into one DataFrame
# forecast_df = pd.concat(forecast_results)

# Ensure that the Excel workbook is saved with all contents
wb.save('Sales_Forecast_Results_transformed.xlsx')
